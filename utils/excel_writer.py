"""
Excel Writer - Utility for creating Excel reports with openpyxl.

Provides reusable functionality for creating formatted Excel workbooks
with multiple sheets, styling, and data validation.
"""

from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelWriter:
    """
    Utility for creating formatted Excel reports.
    
    Provides methods for creating workbooks, adding sheets, styling, and writing data.
    """
    
    def __init__(self, filename: str, output_dir: Path = Path("reports")):
        """Initialize Excel writer."""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.filepath = self.output_dir / filename
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
    
    def add_sheet(self, sheet_name: str) -> 'Sheet':
        """Add a new sheet to workbook."""
        ws = self.workbook.create_sheet(sheet_name)
        return Sheet(ws)
    
    def save(self) -> Path:
        """Save workbook and return file path."""
        self.workbook.save(self.filepath)
        return self.filepath


class Sheet:
    """Wrapper for worksheet with formatting utilities."""
    
    def __init__(self, worksheet):
        """Initialize sheet wrapper."""
        self.ws = worksheet
        self.current_row = 1
    
    def add_header_row(self, headers: List[str], 
                       bg_color: str = "1F4E78",
                       font_color: str = "FFFFFF") -> None:
        """Add header row with formatting."""
        for col_idx, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=self.current_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color=font_color, size=11)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        self.current_row += 1
    
    def add_row(self, values: List, alignment: str = "left") -> None:
        """Add data row with borders."""
        for col_idx, value in enumerate(values, start=1):
            cell = self.ws.cell(row=self.current_row, column=col_idx)
            cell.value = value
            
            # Set alignment
            if alignment == "center":
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif alignment == "right":
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Add border
            cell.border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
        
        self.current_row += 1
    
    def add_summary_row(self, label: str, values: List, 
                       bg_color: str = "E7E6E6") -> None:
        """Add summary row with background color."""
        self.ws.cell(row=self.current_row, column=1).value = label
        
        for col_idx in range(1, len(values) + 2):
            cell = self.ws.cell(row=self.current_row, column=col_idx)
            if col_idx > 1:
                cell.value = values[col_idx - 2]
            
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        self.current_row += 1
    
    def set_column_widths(self, widths: Dict[str, int]) -> None:
        """Set column widths. Keys are column letters (e.g., 'A', 'B', 'C')."""
        for col_letter, width in widths.items():
            self.ws.column_dimensions[col_letter].width = width
    
    def freeze_top_row(self) -> None:
        """Freeze the first row."""
        self.ws.freeze_panes = "A2"
    
    def auto_adjust_columns(self, max_width: int = 50) -> None:
        """Auto-adjust column widths based on content."""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, max_width)
            self.ws.column_dimensions[column_letter].width = adjusted_width
